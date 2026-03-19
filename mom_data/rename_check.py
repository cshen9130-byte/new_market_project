import os
import re
import openpyxl
import warnings
from tqdm import tqdm

warnings.filterwarnings("ignore")

base = "03.投顾逐日"
renamed_files = []
renamed_folders = []
errors = []

# Collect all xlsx files first so tqdm can show total count
# Also track first valid date seen per folder for folder-name check
all_files = []
folder_date = {}  # folder -> date string from inside files

for folder in sorted(os.listdir(base)):
    folder_path = os.path.join(base, folder)
    if not os.path.isdir(folder_path):
        continue
    for fname in os.listdir(folder_path):
        if fname.endswith(".xlsx") and not fname.startswith("~$"):
            all_files.append((folder, folder_path, fname))

for folder, folder_path, fname in tqdm(all_files, desc="Checking files", unit="file"):
    fpath = os.path.join(folder_path, fname)
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        if "品种汇总" not in wb.sheetnames:
            errors.append(f"No 品种汇总 sheet: {fpath}")
            wb.close()
            continue
        ws = wb["品种汇总"]
        account = str(ws["D6"].value).strip() if ws["D6"].value is not None else ""
        date = str(ws["I6"].value).strip() if ws["I6"].value is not None else ""
        wb.close()
    except Exception as e:
        errors.append(f"Error reading {fpath}: {e}")
        continue

    # Record first valid date seen for this folder
    if folder not in folder_date and date:
        folder_date[folder] = date

    # Rename file if needed (use current folder_path which may be stale if folder was renamed)
    expected_name = f"核算信息_{account}_{date}_逐日盯市.xlsx"
    if fname != expected_name:
        new_path = os.path.join(folder_path, expected_name)
        os.rename(fpath, new_path)
        renamed_files.append(f"  [{folder}] {fname}  ->  {expected_name}")

# Now check folder names
# Folder pattern: 恒2 <date>核算单  (prefix may vary, but date is the 8-digit block)
folder_pattern = re.compile(r"^(.*?)(\d{8})(核算单)$")

for folder, date in folder_date.items():
    m = folder_pattern.match(folder)
    if not m:
        # If not matching, forcibly rename to 恒2<date>核算单
        expected_folder = f"恒2 {date}核算单"
        old_path = os.path.join(base, folder)
        new_path = os.path.join(base, expected_folder)
        os.rename(old_path, new_path)
        renamed_folders.append(f"  {folder}  ->  {expected_folder}")
        continue
    prefix, folder_date_str, suffix = m.group(1), m.group(2), m.group(3)
    if folder_date_str != date:
        expected_folder = f"{prefix}{date}{suffix}"
        old_path = os.path.join(base, folder)
        new_path = os.path.join(base, expected_folder)
        os.rename(old_path, new_path)
        renamed_folders.append(f"  {folder}  ->  {expected_folder}")

print(f"\n=== Renamed {len(renamed_files)} file(s) ===")
for r in renamed_files:
    print(r)

print(f"\n=== Renamed {len(renamed_folders)} folder(s) ===")
for r in renamed_folders:
    print(r)

if errors:
    print(f"\n=== Errors ({len(errors)}) ===")
    for e in errors:
        print(e)

if not renamed_files and not renamed_folders and not errors:
    print("All filenames and folder names already match. Nothing to rename.")
